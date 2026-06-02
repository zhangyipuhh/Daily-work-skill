import argparse
import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DEFAULT_INPUT = r'd:\项目文档\AIAssistive\script\checkCode\output\review_results.json'
DEFAULT_OUTPUT = r'd:\项目文档\AIAssistive\script\checkCode\output\review_results.xlsx'


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="将 review_results.json 转换为结构化 Excel"
    )
    parser.add_argument(
        "--input",
        default=DEFAULT_INPUT,
        help=f"输入 JSON 路径（默认：{DEFAULT_INPUT}）",
    )
    parser.add_argument(
        "--output",
        default=DEFAULT_OUTPUT,
        help=f"输出 xlsx 路径（默认：{DEFAULT_OUTPUT}）",
    )
    return parser.parse_args()


args = parse_args()
input_file = args.input
output_file = args.output

if not Path(input_file).exists():
    print(f"[ERROR] 输入文件不存在: {input_file}", file=sys.stderr)
    sys.exit(2)

Path(output_file).parent.mkdir(parents=True, exist_ok=True)

with open(input_file, 'r', encoding='utf-8') as f:
    data = json.load(f)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '代码评审结果'

headers = [
    '开发者', '评审时间',
    '文档质量_总分', '文档质量_完整性_评分', '文档质量_完整性_有需求', '文档质量_完整性_有设计',
    '文档质量_完整性_有实现', '文档质量_完整性_有解决方案', '文档质量_完整性_有总结',
    '文档质量_清晰度_评分', '文档质量_技术准确性_评分',
    '文档质量_AI痕迹_估算比例', '文档质量_AI痕迹_整合质量', '文档质量_AI痕迹_分析',
    'AI采纳率', '数据组合类型',
    '文档代码一致性_评分', '文档代码一致性_有文档', '文档代码一致性_有代码',
    '文档代码一致性_功能匹配', '文档代码一致性_逻辑匹配', '文档代码一致性_接口匹配',
    'AI辅助质量_文档AI比例', 'AI辅助质量_代码AI比例', 'AI辅助质量_有效性',
    '工作流合规_文档优先', '工作流合规_同步开发', '工作流合规_真实反映',
    '文档任务一致性_评分', '文档任务一致性_有任务列表', '文档任务一致性_目标对齐',
    '文档任务一致性_覆盖完整性', '文档任务一致性_范围控制',
    '反模式_存在问题', '反模式_单一功能重复提交', '反模式_与AI编码无关', '反模式_其他问题',
    '改进建议_文档结构', '改进建议_内容完整性', '改进建议_反模式修正',
    '改进建议_AI采纳优化', '改进建议_AI使用优化', '改进建议_优先级',
    '综合评分', '总结', '教练备注'
]

header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

def get_nested(data, *keys, default=''):
    try:
        for key in keys:
            if data is None:
                return default
            data = data[key]
        if data is None:
            return default
        return data
    except (KeyError, TypeError, IndexError):
        return default

def list_to_string(items):
    if not items:
        return ''
    if isinstance(items, list):
        return '; '.join(str(item) for item in items)
    return str(items)

for row_idx, item in enumerate(data, 2):
    d = item.get('data', {})

    row_data = [
        get_nested(d, 'name'),
        get_nested(d, 'review_time'),
        get_nested(d, 'document_quality', 'overall_score'),
        get_nested(d, 'document_quality', 'completeness', 'score'),
        get_nested(d, 'document_quality', 'completeness', 'has_requirement'),
        get_nested(d, 'document_quality', 'completeness', 'has_design'),
        get_nested(d, 'document_quality', 'completeness', 'has_implementation'),
        get_nested(d, 'document_quality', 'completeness', 'has_problem_solution'),
        get_nested(d, 'document_quality', 'completeness', 'has_summary'),
        get_nested(d, 'document_quality', 'clarity', 'score'),
        get_nested(d, 'document_quality', 'technical_accuracy', 'score'),
        get_nested(d, 'document_quality', 'ai_assistance_traces', 'estimated_ai_ratio'),
        get_nested(d, 'document_quality', 'ai_assistance_traces', 'integration_quality'),
        get_nested(d, 'document_quality', 'ai_assistance_traces', 'analysis'),
        get_nested(d, 'ai_adoption_rate', 'adoption_rate'),
        get_nested(d, 'ai_adoption_rate', 'data_combination_type'),
        get_nested(d, 'ai_adoption_rate', 'doc_code_consistency', 'consistency_score'),
        get_nested(d, 'ai_adoption_rate', 'doc_code_consistency', 'has_document'),
        get_nested(d, 'ai_adoption_rate', 'doc_code_consistency', 'has_code'),
        list_to_string(get_nested(d, 'ai_adoption_rate', 'doc_code_consistency', 'function_match')),
        list_to_string(get_nested(d, 'ai_adoption_rate', 'doc_code_consistency', 'logic_match')),
        list_to_string(get_nested(d, 'ai_adoption_rate', 'doc_code_consistency', 'interface_match')),
        get_nested(d, 'ai_adoption_rate', 'ai_assistance_quality', 'doc_ai_ratio'),
        get_nested(d, 'ai_adoption_rate', 'ai_assistance_quality', 'code_ai_ratio'),
        list_to_string(get_nested(d, 'ai_adoption_rate', 'ai_assistance_quality', 'effectiveness')),
        get_nested(d, 'ai_adoption_rate', 'workflow_compliance', 'doc_first'),
        get_nested(d, 'ai_adoption_rate', 'workflow_compliance', 'sync_development'),
        list_to_string(get_nested(d, 'ai_adoption_rate', 'workflow_compliance', 'authentic_reflection')),
        get_nested(d, 'doc_task_consistency', 'score'),
        get_nested(d, 'doc_task_consistency', 'has_task_list'),
        list_to_string(get_nested(d, 'doc_task_consistency', 'goal_alignment')),
        list_to_string(get_nested(d, 'doc_task_consistency', 'coverage_completeness')),
        list_to_string(get_nested(d, 'doc_task_consistency', 'scope_control')),
        get_nested(d, 'anti_patterns', 'has_issues'),
        list_to_string(get_nested(d, 'anti_patterns', 'single_function_repeated_commits', 'analysis')),
        list_to_string(get_nested(d, 'anti_patterns', 'unrelated_to_ai_coding', 'indicators')),
        list_to_string([p['analysis'] for p in get_nested(d, 'anti_patterns', 'other_patterns', default=[]) if isinstance(p, dict)]),
        list_to_string(get_nested(d, 'improvement_suggestions', 'document_structure', default=[])),
        list_to_string(get_nested(d, 'improvement_suggestions', 'content_completeness', default=[])),
        list_to_string(get_nested(d, 'improvement_suggestions', 'anti_pattern_correction', default=[])),
        list_to_string(get_nested(d, 'improvement_suggestions', 'ai_adoption_optimization', default=[])),
        list_to_string(get_nested(d, 'improvement_suggestions', 'ai_usage_optimization', default=[])),
        get_nested(d, 'improvement_suggestions', 'priority'),
        get_nested(d, 'overall_score'),
        get_nested(d, 'summary'),
        get_nested(d, 'coach_notes'),
    ]

    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = thin_border

for col in range(1, len(headers) + 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 20
ws.row_dimensions[1].height = 40

wb.save(output_file)
print(f'Excel文件已生成: {output_file}')
