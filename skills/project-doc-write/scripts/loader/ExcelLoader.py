#!/usr/bin/python
# -*- coding:utf-8 -*-
"""
ExcelLoader 模块（project-doc-write skill 自带）

Date: 2026-06-10
Author: project-doc skill 套件
"""

from pathlib import Path
from typing import List, Dict, Any, Union, Optional

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None  # type: ignore

try:
    from langchain_core.documents import Document  # type: ignore
except Exception:
    Document = None  # type: ignore


class _SimpleDoc:
    __slots__ = ("page_content", "metadata")

    def __init__(self, page_content: str = "", metadata=None):
        self.page_content = page_content
        self.metadata = metadata or {}


class ExcelLoader:
    LOADER_NAME = "ExcelLoader"

    def __init__(self, path: Union[str, Path], data_only: bool = True, include_hidden: bool = False):
        self.path = Path(path)
        self.data_only = data_only
        self.include_hidden = include_hidden
        if load_workbook is None:
            raise ImportError("缺少依赖 openpyxl，请先安装：pip install openpyxl")
        if not self.path.exists():
            raise FileNotFoundError(f"文件不存在: {self.path}")

    def _render_sheet(self, ws) -> str:
        lines: List[str] = [f"## {ws.title}", f"维度: {ws.max_row} 行 x {ws.max_column} 列", ""]
        for row in ws.iter_rows(values_only=True):
            rendered = ["" if v is None else str(v) for v in row]
            lines.append("\t".join(rendered))
        return "\n".join(lines)

    def _load_metadata(self) -> Dict[str, Any]:
        return {
            "source": str(self.path),
            "file_type": self.path.suffix.lower(),
            "loader_used": self.LOADER_NAME,
        }

    def load(self) -> List[Any]:
        wb = load_workbook(filename=str(self.path), data_only=self.data_only, read_only=False)
        docs: List[Any] = []
        base_meta = self._load_metadata()
        for ws in wb.worksheets:
            state = getattr(ws, "sheet_state", "visible")
            if not self.include_hidden and state != "visible":
                continue
            page_content = self._render_sheet(ws)
            sheet_meta = dict(base_meta)
            sheet_meta.update({
                "sheet_name": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "sheet_state": state,
            })
            DocCls = Document if Document is not None else _SimpleDoc
            docs.append(DocCls(page_content=page_content, metadata=sheet_meta))
        return docs

    @property
    def exists(self) -> bool:
        return self.path.exists()
