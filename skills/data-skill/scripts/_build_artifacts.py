#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
构建期生成器：从 数据库规范.xlsx 一次性产出:
  - tables/<sheet_code>.md      (20 个)
  - scripts/validators/<sheet_code>.py (20 个)
运行时不再读 xlsx。
"""
import json
import os
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]  # skill 包根目录 (data-skill/)
XLSX = Path(r"D:\项目文档\AIAssistive\data_skill_tmp\数据库规范.xlsx")
TABLES_DIR = ROOT / "tables"
VAL_DIR = ROOT / "scripts" / "validators"


def split_sheet_name(raw_name: str):
    """
    表名形如: "ZZ_PFXX（农用地转用和土地征收-批复信息表）"
    返回: (code, full_business_name)
      code = "ZZ_PFXX"
      full_business_name = "农用地转用和土地征收-批复信息表"
    """
    m = re.match(r"^([^(（]+)[（(](.+)[)）]\s*$", raw_name)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return raw_name.strip(), raw_name.strip()


def normalize_type(raw_type: str) -> str:
    if raw_type is None:
        return "Char"
    t = raw_type.strip().lower()
    if t in ("char", "clob", "text", "varchar"):
        return "Char"
    if t in ("int", "integer", "bigint"):
        return "INT"
    if t in ("float", "double", "decimal", "number"):
        return "Float"
    if t in ("date", "datetime", "timestamp"):
        return "Date"
    return "Char"


def parse_sheet_rows(rows):
    """
    rows: 二维数组, rows[0] 是表头
    返回: list[dict], 每项包含
      seq, name, code, type, length, decimal, constraint, comment
    重复的字段代码仅保留首次出现。
    """
    out = []
    seen_codes = set()
    for row in rows[1:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        seq, name, code, ftype, flen, fdec, constraint, comment = (
            (list(row) + [None] * 8)[:8]
        )
        if not code or not str(code).strip():
            continue
        code_str = str(code).strip()
        if code_str in seen_codes:
            continue
        seen_codes.add(code_str)
        out.append({
            "seq": str(seq).strip() if seq else "",
            "name": str(name).strip() if name else "",
            "code": code_str,
            "type": normalize_type(ftype),
            "length": int(flen) if flen and str(flen).strip().isdigit() else None,
            "decimal": int(fdec) if fdec and str(fdec).strip().isdigit() else None,
            "constraint": (str(constraint).strip().upper() if constraint else "O"),
            "comment": (str(comment).strip() if comment else ""),
        })
    return out


def render_markdown(code: str, business_name: str, fields: list[dict]) -> str:
    head = (
        f"# {code} - {business_name}\n\n"
        f"- **表代码**: `{code}`\n"
        f"- **业务名称**: {business_name}\n"
        f"- **字段总数**: {len(fields)}\n"
        f"- **主键**: `{fields[0]['code']}`（序号 1，约束 M）\n\n"
        "## 字段说明\n\n"
        "| 序号 | 字段名称 | 字段代码 | 类型 | 长度 | 小数 | 约束 | 备注 |\n"
        "|------|---------|---------|------|------|------|------|------|\n"
    )
    body_rows = []
    for f in fields:
        body_rows.append(
            "| {seq} | {name} | `{code}` | {type} | {length} | {decimal} | {constraint} | {comment} |".format(
                seq=f["seq"],
                name=f["name"],
                code=f["code"],
                type=f["type"],
                length=f["length"] if f["length"] is not None else "",
                decimal=f["decimal"] if f["decimal"] is not None else "",
                constraint=f["constraint"],
                comment=f["comment"] or "",
            )
        )
    return head + "\n".join(body_rows) + "\n"


def build_python_fields_literal(fields: list[dict]) -> str:
    """生成合法的 Python 字面量 (dict list), None/null 正确处理"""
    import pprint
    return pprint.pformat(fields, width=100, sort_dicts=False)


def render_validator(code: str, fields: list[dict]) -> str:
    """
    生成 validate(record) -> (ok, errors)
    校验范围: 类型 / 长度 / 小数位 / 必填(M) / 可空(O)
    """
    lines = [
        "#!/usr/bin/env python",
        "# -*- coding: utf-8 -*-",
        f'"""',
        f"{code} 表的格式校验脚本。",
        f"构建期由 build_artifacts.py 自动生成，请勿手工修改。",
        f"运行时由 SKILL.md 通过 `python {code}.py --record '<json>'` 调用。",
        f'"""',
        "import argparse",
        "import json",
        "import re",
        "import sys",
        "from datetime import datetime",
        "",
        "",
        f"TABLE_CODE = {code!r}",
        f"PRIMARY_KEY = {fields[0]['code']!r}",
        f"FIELDS = {build_python_fields_literal(fields)}",
        "",
        "",
        "def _is_blank(v) -> bool:",
        "    return v is None or (isinstance(v, str) and v.strip() == \"\")",
        "",
        "",
        "def _check_char(value, length, field_code, errors):",
        "    if not isinstance(value, str):",
        "        errors.append(f\"{field_code}: 类型错误, 期望 str, 实际 {type(value).__name__}\")",
        "        return",
        "    if length is not None and len(value) > length:",
        "        errors.append(f\"{field_code}: 长度超限, 最大 {length}, 实际 {len(value)}\")",
        "",
        "",
        "def _check_int(value, field_code, errors):",
        "    if isinstance(value, bool):",
        "        errors.append(f\"{field_code}: 类型错误, 期望 int, 实际 bool\")",
        "        return",
        "    if not isinstance(value, int):",
        "        errors.append(f\"{field_code}: 类型错误, 期望 int, 实际 {type(value).__name__}\")",
        "",
        "",
        "def _check_float(value, decimal, field_code, errors):",
        "    if isinstance(value, bool):",
        "        errors.append(f\"{field_code}: 类型错误, 期望 float, 实际 bool\")",
        "        return",
        "    if not isinstance(value, (int, float)):",
        "        errors.append(f\"{field_code}: 类型错误, 期望 float, 实际 {type(value).__name__}\")",
        "        return",
        "    if decimal is not None:",
        "        s = str(value)",
        "        if \".\" in s:",
        "            frac = s.split(\".\", 1)[1]",
        "            if len(frac) > decimal:",
        "                errors.append(f\"{field_code}: 小数位超限, 最大 {decimal}, 实际 {len(frac)}\")",
        "",
        "",
        "def _check_date(value, field_code, errors):",
        "    if not isinstance(value, str):",
        "        errors.append(f\"{field_code}: 类型错误, 期望 str(ISO 8601), 实际 {type(value).__name__}\")",
        "        return",
        "    if not re.match(r\"^\\d{4}-\\d{2}-\\d{2}([ T]\\d{2}:\\d{2}:\\d{2})?$\", value):",
        "        errors.append(f\"{field_code}: 日期格式错误, 期望 YYYY-MM-DD 或 ISO 8601, 实际 {value!r}\")",
        "        return",
        "    try:",
        "        datetime.fromisoformat(value.replace(\" \", \"T\"))",
        "    except ValueError:",
        "        errors.append(f\"{field_code}: 日期值非法, {value!r}\")",
        "",
        "",
        "def validate(record: dict) -> tuple[bool, list[str]]:",
        "    \"\"\"格式校验",
        "",
    "    Args:",
        "        record: LLM 抽出的字段 dict",
        "",
    "    Returns:",
        "        (ok, errors) 元组",
        "    \"\"\"",
        "    errors: list[str] = []",
        "    if not isinstance(record, dict):",
        "        return False, [\"record 必须是 dict\"]",
        "",
        "    for f in FIELDS:",
        "        code = f[\"code\"]",
        "        ftype = f[\"type\"]",
        "        flen = f[\"length\"]",
        "        fdec = f[\"decimal\"]",
        "        constraint = f[\"constraint\"]",
        "        value = record.get(code)",
        "",
        "        if _is_blank(value):",
        "            if constraint == \"M\":",
        "                errors.append(f\"{code}: 必填字段为空\")",
        "            continue",
        "",
        "        if ftype == \"Char\":",
        "            _check_char(value, flen, code, errors)",
        "        elif ftype == \"INT\":",
        "            _check_int(value, code, errors)",
        "        elif ftype == \"Float\":",
        "            _check_float(value, fdec, code, errors)",
        "        elif ftype == \"Date\":",
        "            _check_date(value, code, errors)",
        "",
        "    return (len(errors) == 0, errors)",
        "",
        "",
        "def main():",
        "    parser = argparse.ArgumentParser(description=f\"{TABLE_CODE} 表格式校验\")",
        "    parser.add_argument(\"--record\", required=True, help=\"JSON 字符串或 @file 路径\")",
        "    args = parser.parse_args()",
        "",
        "    if args.record.startswith(\"@\"):",
        "        with open(args.record[1:], \"r\", encoding=\"utf-8\") as f:",
        "            record = json.load(f)",
        "    else:",
        "        record = json.loads(args.record)",
        "",
        "    ok, errors = validate(record)",
        "    print(json.dumps({\"ok\": ok, \"errors\": errors}, ensure_ascii=False))",
        "    sys.exit(0 if ok else 1)",
        "",
        "",
        "if __name__ == \"__main__\":",
        "    main()",
        "",
    ]
    return "\n".join(lines)


def main():
    TABLES_DIR.mkdir(parents=True, exist_ok=True)
    VAL_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    print(f"读取: {XLSX}")
    print(f"  共 {len(wb.sheetnames)} 张 sheet")

    for sheet_name in wb.sheetnames:
        code, business_name = split_sheet_name(sheet_name)
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            rows.append([str(c) if c is not None else None for c in row])
        fields = parse_sheet_rows(rows)
        if not fields:
            print(f"  [跳过] {sheet_name}: 无有效字段")
            continue

        md_path = TABLES_DIR / f"{code}.md"
        py_path = VAL_DIR / f"{code}.py"
        md_path.write_text(render_markdown(code, business_name, fields), encoding="utf-8")
        py_path.write_text(render_validator(code, fields), encoding="utf-8")
        print(f"  [生成] {code} ({len(fields)} 字段) -> {md_path.relative_to(ROOT)} | {py_path.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
