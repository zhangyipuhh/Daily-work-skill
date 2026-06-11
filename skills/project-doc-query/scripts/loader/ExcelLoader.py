#!/usr/bin/python
# -*- coding:utf-8 -*-
"""
ExcelLoader 模块（project-doc-query skill 自带）

支持 .xlsx / .xlsm 的智能加载：
- 每个 sheet 一个 Document
- 通过 openpyxl 按 sheet 渲染为纯文本
- langchain_core 可选依赖

Date: 2026-06-10
Author: project-doc skill 套件
"""

from pathlib import Path
from typing import List, Dict, Any, Union, Optional

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None  # type: ignore

try:
    from langchain_core.documents import Document  # type: ignore
except Exception:
    Document = None  # type: ignore


class _SimpleDoc:
    __slots__ = ("page_content", "metadata")

    def __init__(self, page_content: str = "", metadata=None):
        self.page_content = page_content
        self.metadata = metadata or {}


class ExcelLoader:
    LOADER_NAME = "ExcelLoader"

    _SHEET_NAME_KEYWORDS = {
        "评审计划": ("评审", "Review", "Plan", "ReviewPlan"),
        "里程碑": ("里程", "Milestone", "节点"),
        "风险登记册": ("风险", "Risk", "风险清单"),
        "WBS": ("WBS", "工作分解", "任务分解"),
        "汇总损益分析": ("损益", "Cost", "成本汇总"),
        "RACI": ("RACI", "责任矩阵", "责任分配"),
    }

    def __init__(
        self,
        path: Union[str, Path],
        data_only: bool = True,
        include_hidden: bool = False,
        sheet_name: Optional[Union[str, List[str]]] = None,
        keyword: Optional[str] = None,
        row_range: Optional[Union[tuple, str]] = None,
    ):
        self.path = Path(path)
        self.data_only = data_only
        self.include_hidden = include_hidden
        self.sheet_name = sheet_name
        self.keyword = keyword
        self.row_range = self._parse_row_range(row_range)
        if load_workbook is None:
            raise ImportError("缺少依赖 openpyxl，请先安装：pip install openpyxl")
        if not self.path.exists():
            raise FileNotFoundError(f"文件不存在: {self.path}")

    @staticmethod
    def _parse_row_range(value):
        if value is None:
            return None
        if isinstance(value, (list, tuple)) and len(value) == 2:
            return (int(value[0]), int(value[1]))
        if isinstance(value, str) and "-" in value:
            parts = value.split("-", 1)
            try:
                return (int(parts[0].strip()), int(parts[1].strip()))
            except ValueError:
                return None
        return None

    @classmethod
    def _match_sheet_name(cls, target: str, available: List[str]) -> List[str]:
        if not target:
            return available
        candidates = [s for s in available if s == target]
        if candidates:
            return candidates
        keywords = []
        for standard, kws in cls._SHEET_NAME_KEYWORDS.items():
            if target in standard or any(kw in target for kw in kws):
                keywords.extend(kws)
                keywords.append(standard)
        if not keywords:
            keywords = [target]
        candidates = [s for s in available if any(kw.lower() in s.lower() for kw in keywords)]
        return candidates

    def _filter_rows(self, ws, rendered_text: str) -> str:
        if not self.keyword and not self.row_range:
            return rendered_text
        lines = rendered_text.split("\n")
        header_lines = [ln for ln in lines if ln.startswith("## ") or ln.startswith("维度:") or ln == ""]
        body_lines = [ln for ln in lines if ln not in header_lines]
        if self.row_range:
            start, end = self.row_range
            body_lines = body_lines[start - 1:end] if body_lines else body_lines
        if self.keyword:
            kw = self.keyword.lower()
            body_lines = [ln for ln in body_lines if kw in ln.lower()]
        return "\n".join(header_lines + body_lines)

    def _render_sheet(self, ws) -> str:
        lines: List[str] = [f"## {ws.title}", f"维度: {ws.max_row} 行 x {ws.max_column} 列", ""]
        for row in ws.iter_rows(values_only=True):
            rendered = ["" if v is None else str(v) for v in row]
            lines.append("\t".join(rendered))
        return "\n".join(lines)

    def _load_metadata(self) -> Dict[str, Any]:
        return {
            "source": str(self.path),
            "file_type": self.path.suffix.lower(),
            "loader_used": self.LOADER_NAME,
        }

    def load(self) -> List[Any]:
        wb = load_workbook(filename=str(self.path), data_only=self.data_only, read_only=False)
        docs: List[Any] = []
        base_meta = self._load_metadata()

        all_sheet_titles = [ws.title for ws in wb.worksheets]

        if self.sheet_name:
            targets = self.sheet_name if isinstance(self.sheet_name, list) else [self.sheet_name]
            matched: List[str] = []
            for t in targets:
                matched.extend(self._match_sheet_name(t, all_sheet_titles))
            matched = list(dict.fromkeys(matched))
            if not matched:
                raise ValueError(
                    f"sheet_name={self.sheet_name!r} 未匹配到任何 sheet。"
                    f"可用的 sheet: {all_sheet_titles}"
                )
            selected_titles = set(matched)
        else:
            selected_titles = set(all_sheet_titles)

        for ws in wb.worksheets:
            if ws.title not in selected_titles:
                continue
            state = getattr(ws, "sheet_state", "visible")
            if not self.include_hidden and state != "visible":
                continue
            page_content = self._render_sheet(ws)
            page_content = self._filter_rows(ws, page_content)
            sheet_meta = dict(base_meta)
            sheet_meta.update({
                "sheet_name": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "sheet_state": state,
            })
            if self.keyword:
                sheet_meta["keyword"] = self.keyword
            if self.row_range:
                sheet_meta["row_range"] = self.row_range
            DocCls = Document if Document is not None else _SimpleDoc
            docs.append(DocCls(page_content=page_content, metadata=sheet_meta))
        return docs

    @property
    def exists(self) -> bool:
        return self.path.exists()


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("用法: python ExcelLoader.py <xlsx_or_xlsm_path>")
        sys.exit(1)
    loader = ExcelLoader(sys.argv[1])
    docs = loader.load()
    print(f"\n总计 {len(docs)} 个 sheet\n")
    for idx, doc in enumerate(docs, 1):
        print(f"--- Sheet {idx}: {doc.metadata.get('sheet_name')} ---")
        print(f"维度: {doc.metadata.get('max_row')} 行 x {doc.metadata.get('max_column')} 列")
        print(f"内容预览:\n{doc.page_content[:600]}{'...' if len(doc.page_content) > 600 else ''}\n")
